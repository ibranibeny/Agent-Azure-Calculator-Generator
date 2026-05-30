"""
Enrich server inventory CSV with OS Version and Disk details
from Excel source files (App-to-Server Export and Rightsizing Export).

This script is used for two-file correlation mode where:
- App-to-Server Export provides OS version, environment, SQL detection
- Rightsizing Export (Disks sheet) provides individual disk details

Usage:
  python scripts/enrich_csv.py --app-export <path> --rightsizing <path> --csv <path>
"""
import openpyxl
import csv
import os
import argparse


def build_os_lookup(app_export_path, header_row=4):
    """Build server -> OS version lookup from App-to-Server Export."""
    print(f"Loading App-to-Server Export: {app_export_path}")
    wb = openpyxl.load_workbook(app_export_path, read_only=True, data_only=True)
    ws = wb["App-to-Server List"]

    os_lookup = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        server = row[0]  # Col A: Server
        os_version = row[8] if len(row) > 8 else None  # Col I: Operating System
        if server and server not in os_lookup and os_version:
            os_lookup[server] = os_version
    wb.close()
    print(f"  OS lookup built: {len(os_lookup)} unique servers")
    return os_lookup


def build_disk_lookup(rightsizing_path, header_row=6):
    """Build server -> disk list from Rightsizing Export Disks sheet."""
    print(f"Loading Rightsizing Disks sheet: {rightsizing_path}")
    wb = openpyxl.load_workbook(rightsizing_path, read_only=True, data_only=True)
    ws = wb["Disks"]

    disk_lookup = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        server = row[0]  # Col A: Server
        if not server:
            continue
        disk_name = row[6] if len(row) > 6 else None  # Col G
        disk_size = row[7] if len(row) > 7 else None  # Col H
        disk_read_iops = row[10] if len(row) > 10 else None  # Col K
        disk_write_iops = row[11] if len(row) > 11 else None  # Col L

        if server not in disk_lookup:
            disk_lookup[server] = []
        disk_lookup[server].append({
            'name': disk_name or '',
            'size_gb': disk_size or 0,
            'read_iops': disk_read_iops or 0,
            'write_iops': disk_write_iops or 0,
        })
    wb.close()
    print(f"  Disk lookup built: {len(disk_lookup)} servers with disk data")
    return disk_lookup


def enrich_csv(input_csv, output_csv, os_lookup, disk_lookup):
    """Read CSV, enrich with OS and disk data, write output."""
    print(f"Reading CSV: {input_csv}")
    rows = []
    with open(input_csv, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        original_fields = list(reader.fieldnames)
        for row in reader:
            rows.append(row)
    print(f"  Read {len(rows)} server rows")

    new_fields = original_fields + ['OSVersion', 'DiskCount', 'TotalDiskGB', 'MaxIOPS', 'DiskDetails']

    print("Enriching with OS and Disk data...")
    os_found = 0
    disk_found = 0

    for row in rows:
        server = row.get('Server', row.get('server', ''))

        # Add OS Version
        os_ver = os_lookup.get(server, '')
        row['OSVersion'] = os_ver
        if os_ver:
            os_found += 1

        # Add Disk info
        disks = disk_lookup.get(server, [])
        row['DiskCount'] = len(disks)
        if disks:
            disk_found += 1
            total_size = sum(float(d['size_gb']) for d in disks)
            max_iops = max(float(d['read_iops']) + float(d['write_iops']) for d in disks)
            detail_parts = [f"{d['name']}:{d['size_gb']}GB" for d in disks]
            row['TotalDiskGB'] = round(total_size, 2)
            row['MaxIOPS'] = round(max_iops, 2)
            row['DiskDetails'] = '; '.join(detail_parts)
        else:
            row['TotalDiskGB'] = 0
            row['MaxIOPS'] = 0
            row['DiskDetails'] = ''

    # Write enriched CSV
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=new_fields)
        writer.writeheader()
        writer.writerows(rows)

    print(f"✅ Enriched CSV written: {output_csv}")
    print(f"   OS matches: {os_found}/{len(rows)}")
    print(f"   Disk matches: {disk_found}/{len(rows)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Enrich server CSV with OS and Disk data from Excel")
    parser.add_argument("--app-export", required=True, help="Path to App-to-Server Export Excel file")
    parser.add_argument("--rightsizing", required=True, help="Path to Rightsizing Export Excel file")
    parser.add_argument("--csv", required=True, help="Path to server inventory CSV to enrich")
    parser.add_argument("--output", help="Output CSV path (defaults to overwriting input)")
    args = parser.parse_args()

    output_path = args.output or args.csv

    os_lookup = build_os_lookup(args.app_export)
    disk_lookup = build_disk_lookup(args.rightsizing)
    enrich_csv(args.csv, output_path, os_lookup, disk_lookup)
